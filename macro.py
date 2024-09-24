import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from datetime import datetime
import requests
import base64
import configparser
import os
import json
import asyncio
import threading
import time


config = configparser.ConfigParser()
config.read("config.ini")
search_url = "https://console.capellaspace.com/search/"
df = None
sheet_buttons = []
date = None
capella = None
access_token = None
output_folder_label = None
collect_id_button = None
download_button = None
selected_sheet_name = None
selected_date = None


def open_excel():
    global df
    filepath = filedialog.askopenfilename(
        initialdir="/",
        title="엑셀 파일 선택",
        filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*")),
    )

    if filepath:
        if read_excel(filepath):
            update_sheet_buttons()


def update_sheet_buttons():
    global df, sheet_buttons

    for button in sheet_buttons:
        button.destroy()
    sheet_buttons = []

    if df is not None:
        for sheet_name in df.sheetnames:
            button = tk.Button(
                sheet_frame,
                text=sheet_name,
                command=lambda name=sheet_name: on_sheet_button_click(name),
            )
            button.pack(side=tk.TOP, anchor=tk.W, fill=tk.X)
            sheet_buttons.append(button)


def on_sheet_button_click(name):
    global df, date, capella, collect_id, selected_sheet_name
    selected_sheet_name = name
    sheet = df[selected_sheet_name]
    selected_sheet_label.config(text=f"선택한 시트: {selected_sheet_name}")

    filtered_data = []
    previous_date = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if isinstance(row[1], datetime):
            formatted_date = row[1].strftime("%Y-%m-%d")
        else:
            formatted_date = row[1]

        if formatted_date is None:
            formatted_date = previous_date

        filtered_data.append(
            {
                "날짜": formatted_date,
                "Capella": row[2],
                "collect_id": row[3],
            }
        )

        previous_date = formatted_date

    update_filtered_buttons(filtered_data)


def update_filtered_buttons(filtered_data):
    global sheet_buttons

    for button in sheet_buttons:
        button.destroy()
    sheet_buttons = []

    scrollbar = tk.Scrollbar(sheet_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    canvas = tk.Canvas(sheet_frame, yscrollcommand=scrollbar.set)
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=canvas.yview)

    button_frame = tk.Frame(canvas)
    canvas.create_window((0, 0), window=button_frame, anchor="nw")

    for item in filtered_data:
        button_text = f"{item['날짜']} - {item['Capella']}"
        button = tk.Button(
            button_frame,
            text=button_text,
            command=lambda item=item: on_filtered_button_click(item),
        )
        button.pack(side=tk.TOP, anchor=tk.W, fill=tk.X)
        sheet_buttons.append(button)

    # Update the scroll region
    button_frame.update_idletasks()
    canvas.config(scrollregion=canvas.bbox("all"))


def on_filtered_button_click(item):
    global collect_id_button, download_button, selected_date
    collect_id = item["collect_id"]
    if collect_id:
        download_button_text = f"다운로드하기 ({collect_id})"

        if collect_id_button is not None:
            collect_id_button.destroy()
        if download_button is not None:
            download_button.destroy()

        selected_date = item["날짜"]

        download_button = tk.Button(
            main_frame,
            text=download_button_text,
            command=lambda: download_archive_capella(item),
        )
        download_button.pack(pady=5)


def read_excel(filepath):
    global df
    try:
        df = openpyxl.load_workbook(filepath)
        print("엑셀 파일이 성공적으로 열렸습니다.")
        return True
    except Exception as e:
        print(f"엑셀 파일을 여는 중 오류 발생: {e}")
        return False


def login_success():
    global username_label, username_entry, password_label, password_entry, submit_button

    username_label.pack_forget()
    username_entry.pack_forget()
    password_label.pack_forget()
    password_entry.pack_forget()
    submit_button.pack_forget()

    success_label = tk.Label(main_frame, text="로그인 성공")
    success_label.pack(pady=20)


def download_archive_capella(item):
    global outputdir, selected_sheet_name
    """영상 다운로드"""

    if outputdir is None:
        print("출력 디렉토리가 설정되지 않았습니다.")
        return

    if selected_sheet_name is None:
        print("시트 이름이 설정되지 않았습니다.")
        return

    accesstoken = config["Bot"]["token"]
    download_type = {
        key.upper() for key, value in config["Download"].items() if value == "True"
    }
    print(f"Download Type : {download_type}")
    print(f"Start : {item}")
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/geo+json",
        "Authorization": f"Bearer {accesstoken}",
    }
    search_url = "https://api.capellaspace.com/catalog/search"

    if accesstoken:
        try:
            result = requests.post(
                search_url,
                headers=headers,
                json={"query": {"capella:collect_id": {"eq": item["collect_id"]}}},
            ).json()
            download_url = "https://api.capellaspace.com/orders"
            # print(json.dumps(result, indent=4))

            for feature in result["features"]:
                # print("product_type : ", feature["properties"]["sar:product_type"])
                # print("collection : ", feature["collection"])
                # print("id : ", feature["id"])

                if feature["properties"]["sar:product_type"] in download_type:
                    try:
                        if feature["collection"] == "capella-beta-analytics":
                            response = requests.post(
                                download_url,
                                headers=headers,
                                json={
                                    "items": [
                                        {
                                            "collectionId": feature["collection"],
                                            "granuleId": feature["id"],
                                        },
                                        {
                                            "collectionId": "capella-geo",
                                            "granuleId": item["Capella"],
                                        },
                                    ]
                                },
                            ).json()
                        else:
                            response = requests.post(
                                download_url,
                                headers=headers,
                                json={
                                    "items": [
                                        {
                                            "collectionId": feature["collection"],
                                            "granuleId": feature["id"],
                                        }
                                    ]
                                },
                            ).json()

                        # print(response)
                        order_id = response["orderId"]
                        print(f"Order ID : {order_id}")
                        id = feature["id"]
                        print("id : ", id)

                        download_url = f"{download_url}/{order_id}/download"
                        response = requests.get(
                            download_url,
                            headers=headers,
                            json={
                                "items": [
                                    {
                                        "collectionId": feature["collection"],
                                        "granuleId": [feature["id"]],
                                    }
                                ]
                            },
                        )
                        jResult = response.json()
                        if not response.status_code in [200, 201, 202]:
                            error_data = response.json()["error"]["message"]
                            print(f"Error: {error_data}")
                            print(f"Error: {json.dumps(response.json(), indent=4)}")

                        for info in jResult:
                            if feature["collection"] == "capella-beta-analytics":
                                if "GEO" in info["id"]:
                                    continue
                            save_id = info["id"]
                            asset_keys = info["assets"].keys()
                            asset_values = info["assets"]

                            save_path = os.path.join(
                                outputdir, selected_sheet_name, item["날짜"], id
                            )
                            if os.path.exists(save_path):
                                continue
                            os.makedirs(save_path, exist_ok=True)

                            for key in asset_keys:
                                if key != "license":
                                    signed_url = asset_values[key]["hrefDownload"]
                                    filename = signed_url[signed_url.rfind("/") + 1 :]
                                    sep = "?"
                                    downloadfilename = filename.split(sep, 1)[0]
                                    if save_id in signed_url:
                                        try:
                                            with requests.get(signed_url) as result:
                                                result.raise_for_status()
                                                dlFileName = os.path.join(
                                                    save_path, downloadfilename
                                                )
                                                with open(dlFileName, "wb") as f:
                                                    for chunk in result.iter_content(
                                                        chunk_size=1024
                                                    ):
                                                        f.write(chunk)
                                            print(
                                                f"Download Complete : {downloadfilename}"
                                            )
                                        except (
                                            requests.exceptions.RequestException
                                        ) as e:
                                            print(f"Error : {e}")
                    except requests.exceptions.RequestException as e:
                        print(f"Request error: {e}")
        except requests.exceptions.RequestException as e:
            print(f"Request error: {e}")


def submit_credentials():
    global username_entry, password_entry, access_token

    username = username_entry.get()
    password = password_entry.get()

    if "." not in config["Bot"]["token"]:
        # 토큰 발급 요청
        token_url = "https://api.capellaspace.com/token"
        auth_header = base64.b64encode(f"{username}:{password}".encode()).decode()
        headers = {
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {auth_header}",
        }
        data = {"grant_type": "client_credentials"}

        try:
            response = requests.post(token_url, headers=headers, data=data)
            response.raise_for_status()

            token_data = response.json()
            access_token = token_data["accessToken"]
            print(f"액세스 토큰: {access_token}")

            config["Bot"]["token"] = access_token
            config["account"]["username"] = username
            config["account"]["password"] = password
            with open("config.ini", "w") as configfile:
                config.write(configfile)

            login_success()

        except requests.exceptions.RequestException as e:
            print(f"토큰 발급 오류: {e}")
    else:
        login_success()


def choose_output_folder():
    global outputdir, output_folder_label

    outputdir = filedialog.askdirectory(initialdir="/", title="출력 폴더 선택")
    if outputdir:
        print(f"선택된 출력 폴더: {outputdir}")

        if output_folder_label is not None:
            output_folder_label.destroy()

        output_folder_label = tk.Label(main_frame, text=f"출력 폴더: {outputdir}")
        output_folder_label.pack(pady=5)


def download_all_sheets():
    sheet = df[selected_sheet_name]
    previous_date = None
    tasks = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        collect_id = row[3]
        if collect_id:
            item = {
                "collect_id": collect_id,
                "날짜": (
                    row[1].strftime("%Y-%m-%d")
                    if isinstance(row[1], datetime)
                    else row[1]
                ),
                "Capella": row[2],
            }
            if item["날짜"] is None:
                item["날짜"] = f"{previous_date}_2"

            print(
                "date: ",
                item["날짜"],
            )
            previous_date = item["날짜"]
            tasks.append(download_archive_capella_async(item))
    loop = asyncio.get_event_loop()
    loop.run_until_complete(asyncio.gather(*tasks))


async def download_archive_capella_async(item):
    await asyncio.to_thread(download_archive_capella, item)


def go_back():
    update_sheet_buttons()
    selected_sheet_label.config(text="선택한 시트: 없음")


def refresh_token():
    id = config["account"]["username"]
    pw = config["account"]["password"]

    if id == "":
        return
    if pw == "":
        return
    token_url = "https://api.capellaspace.com/token"
    auth_header = base64.b64encode(f"{id}:{pw}".encode()).decode()
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "Authorization": f"Basic {auth_header}",
    }
    data = {"grant_type": "client_credentials"}

    try:
        response = requests.post(token_url, headers=headers, data=data)
        response.raise_for_status()

        token_data = response.json()
        access_token = token_data["accessToken"]
        print(f"액세스 토큰: {access_token}")

        config["Bot"]["token"] = access_token
        with open("config.ini", "w") as configfile:
            config.write(configfile)

        login_success()

    except requests.exceptions.RequestException as e:
        pass


def start_refresh_token_thread():
    while True:
        time.sleep(2400)
        refresh_token()


def update_config(option, value):
    config["Download"][option] = str(value)
    with open("config.ini", "w") as configfile:
        config.write(configfile)


def create_checkboxes():
    global checkboxes
    checkboxes = {}

    for option in ["geo", "gec", "sicd", "slc", "ar"]:
        var = tk.BooleanVar(value=config.getboolean("Download", option))
        checkbox = tk.Checkbutton(
            main_frame,
            text=option,
            variable=var,
            command=lambda opt=option, v=var: update_config(opt, v.get()),
        )
        checkbox.pack(anchor=tk.W)
        checkboxes[option] = var


window = tk.Tk()
window.title("Capella Downloader")

sheet_frame = tk.Frame(window, width=200)
sheet_frame.pack(side=tk.LEFT, fill=tk.Y)

main_frame = tk.Frame(window)
main_frame.pack(side=tk.LEFT, padx=20, pady=20)

refresh_token_button = tk.Button(
    main_frame, text="토큰 새로 고침", command=refresh_token
)
refresh_token_button.pack(pady=10)

open_button = tk.Button(main_frame, text="엑셀 파일 열기", command=open_excel)
open_button.pack(pady=10)

selected_sheet_label = tk.Label(main_frame, text="선택한 시트: 없음")
selected_sheet_label.pack(pady=5)

download_all_button = tk.Button(
    main_frame, text="전부 다운로드", command=download_all_sheets
)
download_all_button.pack(pady=10)

back_button = tk.Button(main_frame, text="뒤로가기", command=go_back)
back_button.pack(pady=10)

username_label = tk.Label(main_frame, text="아이디:")
username_label.pack()
username_entry = tk.Entry(main_frame)
username_entry.insert(0, "")
username_entry.pack()


password_label = tk.Label(main_frame, text="비밀번호:")
password_label.pack()
password_entry = tk.Entry(main_frame, show="*")
password_entry.insert(0, "")
password_entry.pack()

submit_button = tk.Button(main_frame, text="제출", command=submit_credentials)
submit_button.pack(pady=10)

choose_folder_button = tk.Button(
    main_frame, text="출력 폴더 선택", command=choose_output_folder
)
choose_folder_button.pack(pady=10)

create_checkboxes()

window.mainloop()
